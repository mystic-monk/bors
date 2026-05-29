from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib.admin.views.decorators import staff_member_required

from decimal import Decimal, InvalidOperation
from .models import OperatorReturn, OperatorAccess, Declaration, VehicleRouteUsage, VehicleEmission, SeedLicence, SeedVehicle
from .forms import GeneralForm, LicenceFormSet, EmissionFormSet, AccessibilityFormSet, OperatorAccessForm, DeclarationForm

STEPS = [
    ('general',       'General',       1),
    ('licence',       'Licences',      2),
    ('emissions',     'Emissions',     3),
    ('accessibility', 'Accessibility', 4),
    ('declaration',   'Declaration',   5),
]


# ── Access guard ─────────────────────────────────────────────────────────────

def require_operator_access(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        access_id = request.session.get('operator_access_id')
        if not access_id:
            return redirect('access_login')
        try:
            access = OperatorAccess.objects.get(pk=access_id, is_active=True)
        except OperatorAccess.DoesNotExist:
            request.session.pop('operator_access_id', None)
            return redirect('access_login')
        request.operator_access = access
        return view_func(request, *args, **kwargs)
    return wrapper


def _owns_return(access, pk):
    """Return the OperatorReturn if it belongs to this access record, else None."""
    if access.operator_return_id == pk:
        return access.operator_return
    return None


# ── Auth views ────────────────────────────────────────────────────────────────

def access_login(request):
    if request.session.get('operator_access_id'):
        return redirect('index')

    error = None
    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        try:
            access = OperatorAccess.objects.get(token=token, is_active=True)
            access.last_accessed_at = timezone.now()
            access.save(update_fields=['last_accessed_at'])
            request.session['operator_access_id'] = access.pk
            return redirect('index')
        except OperatorAccess.DoesNotExist:
            error = 'Invalid or expired access code. Please check your email and try again.'

    return render(request, 'returns/access_login.html', {'error': error})


def access_logout(request):
    request.session.pop('operator_access_id', None)
    return redirect('access_login')


# ── Staff token management ─────────────────────────────────────────────────────

@staff_member_required(login_url='/admin/login/')
def admin_tokens(request):
    # If an operator session is active in this browser, block access entirely.
    if request.session.get('operator_access_id'):
        return redirect('index')

    form = OperatorAccessForm()
    if request.method == 'POST':
        form = OperatorAccessForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'Access token created for {form.cleaned_data["operator_name"]}.')
            return redirect('admin_tokens')

    accesses = OperatorAccess.objects.select_related('operator_return').order_by('-created_at')
    return render(request, 'returns/admin_tokens.html', {
        'accesses': accesses,
        'form': form,
    })


@staff_member_required(login_url='/admin/login/')
def toggle_access(request, pk):
    access = get_object_or_404(OperatorAccess, pk=pk)
    access.is_active = not access.is_active
    access.save(update_fields=['is_active'])
    status = 'activated' if access.is_active else 'deactivated'
    messages.success(request, f'{access.operator_name} access {status}.')
    return redirect('admin_tokens')


@staff_member_required(login_url='/admin/login/')
def regenerate_token(request, pk):
    if request.session.get('operator_access_id') or request.method != 'POST':
        return redirect('index')
    access = get_object_or_404(OperatorAccess, pk=pk)
    access.regenerate_token()
    messages.success(request, f'New token generated for {access.operator_name}.')
    return redirect('admin_tokens')


@staff_member_required(login_url='/admin/login/')
def regenerate_all_tokens(request):
    if request.session.get('operator_access_id') or request.method != 'POST':
        return redirect('index')
    count = 0
    for access in OperatorAccess.objects.all():
        access.regenerate_token()
        count += 1
    messages.success(request, f'All {count} tokens have been regenerated. Send the new tokens to each operator before the next collection period.')
    return redirect('admin_tokens')


# ── Operator views ─────────────────────────────────────────────────────────────

@require_operator_access
def index(request):
    access = request.operator_access
    return render(request, 'returns/index.html', {
        'operator_return': access.operator_return,
        'access': access,
    })


@require_operator_access
def step_general(request, pk=None):
    access = request.operator_access

    if pk:
        # Ensure the pk belongs to this operator
        if not _owns_return(access, pk):
            messages.error(request, 'You do not have permission to edit that return.')
            return redirect('index')
        instance = get_object_or_404(OperatorReturn, pk=pk)
    else:
        # If they already have a return, edit it rather than create a second one
        instance = access.operator_return

    if request.method == 'POST':
        form = GeneralForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save()
            if not access.operator_return:
                access.operator_return = obj
                access.save(update_fields=['operator_return'])
            messages.success(request, 'General details saved.')
            return redirect('step_licence', pk=obj.pk)
    else:
        initial = {}
        if not instance:
            initial['operator_name'] = access.operator_name
            initial['data_year'] = 2025
        form = GeneralForm(instance=instance, initial=initial)

    return render(request, 'returns/step_general.html', {
        'form': form,
        'steps': STEPS,
        'current_step': 1,
        'pk': instance.pk if instance else None,
    })


@require_operator_access
def step_licence(request, pk):
    if not _owns_return(request.operator_access, pk):
        messages.error(request, 'You do not have permission to edit that return.')
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)

    # Apply NTA seed routes on first visit (no licences yet)
    if request.method == 'GET' and not obj.licences.exists():
        access = getattr(obj, 'access', None)
        if access:
            for seed in access.seed_licences.all():
                obj.licences.create(
                    route_no=seed.route_no,
                    on_free_travel=False,
                    on_yasc=False,
                    date_joined=timezone.localdate(),
                    service_commenced=False,
                    service_ceased=False,
                    total_passenger_journeys=0,
                    free_travel_percent=0,
                    km_operated=0,
                    daily_pvr=0,
                )

    if request.method == 'POST':
        formset = LicenceFormSet(request.POST, instance=obj)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Licence details saved.')
            return redirect('step_emissions', pk=pk)
    else:
        formset = LicenceFormSet(instance=obj)

    return render(request, 'returns/step_licence.html', {
        'formset': formset,
        'obj': obj,
        'steps': STEPS,
        'current_step': 2,
        'pk': pk,
    })


@require_operator_access
def step_emissions(request, pk):
    if not _owns_return(request.operator_access, pk):
        messages.error(request, 'You do not have permission to edit that return.')
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)
    routes = list(obj.licences.values_list('route_no', flat=True).order_by('route_no'))

    # Apply NTA seed vehicles on first visit (no emissions yet)
    if request.method == 'GET' and not obj.emissions.exists():
        access = getattr(obj, 'access', None)
        if access:
            for i, seed in enumerate(access.seed_vehicles.all(), start=1):
                obj.emissions.create(
                    res_id=seed.res_id or str(i),
                    vehicle_reg=seed.vehicle_reg,
                    regularly_deployed=False,
                    make=seed.make,
                    model_version=seed.model_version,
                    transmission=seed.transmission,
                    engine_type=seed.engine_type,
                    fuel_type='Diesel',
                    euro_standard='',
                    make_model_correct=True,
                    seats_on_record=seed.seats_on_record,
                    correct_seats=None,
                    avl_gps=False,
                )

    if request.method == 'POST':
        formset = EmissionFormSet(request.POST, instance=obj)
        if formset.is_valid():
            formset.save()
            for form_idx, form in enumerate(formset.forms):
                if form.instance.pk and not form.cleaned_data.get('DELETE'):
                    vehicle = form.instance
                    for route_idx, route_no in enumerate(routes):
                        raw = request.POST.get(f'route_pct_{form_idx}_{route_idx}', '').strip()
                        try:
                            pct = Decimal(raw) if raw else None
                        except InvalidOperation:
                            pct = None
                        VehicleRouteUsage.objects.update_or_create(
                            vehicle=vehicle, route_no=route_no,
                            defaults={'usage_percent': pct},
                        )
            messages.success(request, 'Emissions data saved.')
            return redirect('step_accessibility', pk=pk)
    else:
        formset = EmissionFormSet(instance=obj)

    existing_usages = {}
    for emission in obj.emissions.prefetch_related('route_usages').all():
        existing_usages[emission.pk] = {u.route_no: u.usage_percent for u in emission.route_usages.all()}

    forms_with_usages = []
    for form in formset.forms:
        pk_val = form.instance.pk
        usages = existing_usages.get(pk_val, {})
        route_data = [(route_no, i, usages.get(route_no, '')) for i, route_no in enumerate(routes)]
        forms_with_usages.append((form, route_data))

    return render(request, 'returns/step_emissions.html', {
        'formset': formset,
        'forms_with_usages': forms_with_usages,
        'routes': routes,
        'obj': obj,
        'steps': STEPS,
        'current_step': 3,
        'pk': pk,
    })


@require_operator_access
def step_accessibility(request, pk):
    if not _owns_return(request.operator_access, pk):
        messages.error(request, 'You do not have permission to edit that return.')
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)

    # Ensure one accessibility record exists per emission vehicle
    emissions = list(obj.emissions.all().order_by('id'))
    existing_regs = set(obj.accessibility.values_list('vehicle_reg', flat=True))
    for em in emissions:
        if em.vehicle_reg not in existing_regs:
            obj.accessibility.create(res_id=em.res_id, vehicle_reg=em.vehicle_reg)

    if request.method == 'POST':
        formset = AccessibilityFormSet(request.POST, instance=obj)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Accessibility data saved.')
            return redirect('step_declaration', pk=pk)
    else:
        formset = AccessibilityFormSet(instance=obj)

    emissions_by_reg = {em.vehicle_reg: em for em in emissions}
    forms_with_emissions = [
        (form, emissions_by_reg.get(form.instance.vehicle_reg))
        for form in formset.forms
        if form.instance.vehicle_reg  # skip the extra blank form
    ]

    return render(request, 'returns/step_accessibility.html', {
        'formset': formset,
        'forms_with_emissions': forms_with_emissions,
        'obj': obj,
        'steps': STEPS,
        'current_step': 4,
        'pk': pk,
    })


@require_operator_access
def step_declaration(request, pk):
    if not _owns_return(request.operator_access, pk):
        messages.error(request, 'You do not have permission to edit that return.')
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)

    instance = getattr(obj, 'declaration', None)

    if request.method == 'POST':
        form = DeclarationForm(request.POST, instance=instance)
        if form.is_valid():
            decl = form.save(commit=False)
            decl.operator_return = obj
            decl.save()
            obj.status = 'submitted'
            obj.submitted_at = timezone.now()
            obj.save()
            messages.success(request, 'Return submitted successfully.')
            return redirect('success', pk=pk)
    else:
        initial = {}
        if not instance:
            initial['declaration_date'] = timezone.localdate()
        form = DeclarationForm(instance=instance, initial=initial)

    return render(request, 'returns/step_declaration.html', {
        'form': form,
        'obj': obj,
        'steps': STEPS,
        'current_step': 5,
        'pk': pk,
    })


@require_operator_access
def success(request, pk):
    if not _owns_return(request.operator_access, pk):
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)
    return render(request, 'returns/success.html', {'obj': obj})


@require_operator_access
def export_excel(request, pk):
    if not _owns_return(request.operator_access, pk):
        messages.error(request, 'You do not have permission to export that return.')
        return redirect('index')
    obj = get_object_or_404(OperatorReturn, pk=pk)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    center = Alignment(horizontal='center')

    def style_header_row(ws, row_num):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    ws1 = wb.active
    ws1.title = 'General Questions'
    ws1.append(['Data Year', 'Operator Name', 'Trading Name', 'County', 'Annual Revenue (€)', 'TaxSaver'])
    style_header_row(ws1, 1)
    ws1.append([
        obj.data_year, obj.operator_name, obj.trading_name,
        obj.county, float(obj.annual_revenue),
        obj.get_taxsaver_display() + (f' — {obj.taxsaver_other}' if obj.taxsaver == 'other' else ''),
    ])

    ws2 = wb.create_sheet('LicenceQuestions')
    ws2.append([
        'Operator Name', 'Licence No', 'Route No', 'On Free Travel', 'On YASC',
        'Date Joined', 'Service Commenced', 'Service Ceased', 'Date Ceased',
        'Total Passenger Journeys', 'Free Travel %', 'KM Operated', 'Daily PVR',
    ])
    style_header_row(ws2, 1)
    for lic in obj.licences.all():
        ws2.append([
            obj.operator_name, lic.licence_no, lic.route_no,
            'Y' if lic.on_free_travel else 'N',
            'Y' if lic.on_yasc else 'N',
            lic.date_joined.strftime('%d/%m/%Y') if lic.date_joined else '',
            'Y' if lic.service_commenced else 'N',
            'Y' if lic.service_ceased else 'N',
            lic.date_ceased.strftime('%d/%m/%Y') if lic.date_ceased else '',
            lic.total_passenger_journeys,
            float(lic.free_travel_percent),
            float(lic.km_operated),
            lic.daily_pvr,
        ])

    ws3 = wb.create_sheet('Emissions')
    ws3.append([
        'RES ID', 'Reg No', 'Regularly Deployed', 'Make', 'Model',
        'Transmission', 'Engine Type', 'Fuel Type', 'Euro Standard',
        'Make/Model Correct', 'Seats on Record', 'Correct Seats', 'AVL/GPS',
    ])
    style_header_row(ws3, 1)
    for v in obj.emissions.all():
        ws3.append([
            v.res_id, v.vehicle_reg,
            'Y' if v.regularly_deployed else 'N',
            v.make, v.model_version, v.transmission, v.engine_type,
            v.fuel_type, v.euro_standard,
            'Y' if v.make_model_correct else 'N',
            v.seats_on_record, v.correct_seats,
            'Y' if v.avl_gps else 'N',
        ])

    ws4 = wb.create_sheet('Accessibility Questions')
    ws4.append([
        'RES ID', 'Reg No', 'Wheelchair Spaces', 'Access Type',
        'Mechanical Ramp', 'Centre Door Ramps', 'Exterior Displays',
        'Interior Stop Displays', 'Audio Announcements', 'Yellow Handrails',
        'Priority Seating', 'Contrasting Seat Covers', 'Induction Loop',
    ])
    style_header_row(ws4, 1)
    for a in obj.accessibility.all():
        ws4.append([
            a.res_id, a.vehicle_reg, a.wheelchair_spaces, a.wheelchair_access_type,
            'Y' if a.mechanical_ramp else 'N',
            'Y' if a.centre_door_ramps else 'N',
            'Y' if a.exterior_displays else 'N',
            'Y' if a.interior_stop_displays else 'N',
            'Y' if a.audio_announcements else 'N',
            'Y' if a.yellow_handrails else 'N',
            'Y' if a.priority_seating else 'N',
            'Y' if a.contrasting_seat_covers else 'N',
            'Y' if a.induction_loop else 'N',
        ])

    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'BORS_{obj.operator_name.replace(" ", "_")}_{obj.data_year}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Admin bulk exports ────────────────────────────────────────────────────────

@staff_member_required(login_url='/admin/login/')
def export_all_excel(request):
    returns = (
        OperatorReturn.objects
        .prefetch_related('licences', 'emissions__route_usages', 'accessibility', 'declaration')
        .order_by('operator_name', 'data_year')
    )

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    ctr = Alignment(horizontal='center')

    def style_header(ws):
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ctr

    def autowidth(ws):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

    # Sheet 1 — General
    ws1 = wb.active
    ws1.title = 'General'
    ws1.append(['Operator Name', 'Trading Name', 'Data Year', 'County', 'Annual Revenue (€)', 'TaxSaver', 'Status', 'Submitted At'])
    style_header(ws1)
    for r in returns:
        ws1.append([
            r.operator_name, r.trading_name, r.data_year, r.county,
            float(r.annual_revenue),
            r.get_taxsaver_display() + (f' — {r.taxsaver_other}' if r.taxsaver == 'other' else ''),
            r.get_status_display(),
            r.submitted_at.strftime('%d/%m/%Y %H:%M') if r.submitted_at else '',
        ])
    autowidth(ws1)

    # Sheet 2 — Licences
    ws2 = wb.create_sheet('Licences')
    ws2.append([
        'Operator Name', 'Data Year', 'Route No', 'Licence No',
        'On Free Travel', 'On YASC', 'Date Joined YASC',
        'Service Commenced', 'Service Ceased', 'Date Ceased',
        'Total Passenger Journeys', 'Free Travel %', 'KM Operated', 'Daily PVR',
    ])
    style_header(ws2)
    for r in returns:
        for lic in r.licences.all():
            ws2.append([
                r.operator_name, r.data_year, lic.route_no, lic.licence_no,
                'Y' if lic.on_free_travel else 'N',
                'Y' if lic.on_yasc else 'N',
                lic.date_joined.strftime('%d/%m/%Y') if lic.date_joined else '',
                'Y' if lic.service_commenced else 'N',
                'Y' if lic.service_ceased else 'N',
                lic.date_ceased.strftime('%d/%m/%Y') if lic.date_ceased else '',
                lic.total_passenger_journeys,
                float(lic.free_travel_percent),
                float(lic.km_operated),
                lic.daily_pvr,
            ])
    autowidth(ws2)

    # Sheet 3 — Emissions
    ws3 = wb.create_sheet('Emissions')
    ws3.append([
        'Operator Name', 'Data Year', 'RES ID', 'Reg No',
        'Regularly Deployed', 'Make', 'Model Version', 'Transmission', 'Engine Type',
        'Make/Model Correct', 'Fuel Type', 'Euro Standard',
        'Seats on Record', 'Correct Seats', 'AVL/GPS',
    ])
    style_header(ws3)
    for r in returns:
        for v in r.emissions.all():
            ws3.append([
                r.operator_name, r.data_year, v.res_id, v.vehicle_reg,
                'Y' if v.regularly_deployed else 'N',
                v.make, v.model_version, v.transmission, v.engine_type,
                'Y' if v.make_model_correct else 'N',
                v.fuel_type, v.euro_standard,
                v.seats_on_record, v.correct_seats,
                'Y' if v.avl_gps else 'N',
            ])
    autowidth(ws3)

    # Sheet 4 — Vehicle Route Usage
    ws4 = wb.create_sheet('Vehicle Route Usage')
    ws4.append(['Operator Name', 'Data Year', 'Vehicle Reg No', 'Route No', 'Usage %'])
    style_header(ws4)
    for r in returns:
        for v in r.emissions.all():
            for u in v.route_usages.all():
                ws4.append([r.operator_name, r.data_year, v.vehicle_reg, u.route_no,
                             float(u.usage_percent) if u.usage_percent is not None else ''])
    autowidth(ws4)

    # Sheet 5 — Accessibility
    ws5 = wb.create_sheet('Accessibility')
    ws5.append([
        'Operator Name', 'Data Year', 'RES ID', 'Reg No',
        'Wheelchair Spaces', 'Wheelchair Access Type',
        'Mechanical Ramp', 'Centre Door Ramps', 'Exterior Displays',
        'Interior Stop Displays', 'Audio Announcements', 'Yellow Handrails',
        'Priority Seating', 'Contrasting Seat Covers', 'Induction Loop',
    ])
    style_header(ws5)
    for r in returns:
        for a in r.accessibility.all():
            ws5.append([
                r.operator_name, r.data_year, a.res_id, a.vehicle_reg,
                a.wheelchair_spaces or 0, a.wheelchair_access_type,
                'Y' if a.mechanical_ramp else 'N',
                'Y' if a.centre_door_ramps else 'N',
                'Y' if a.exterior_displays else 'N',
                'Y' if a.interior_stop_displays else 'N',
                'Y' if a.audio_announcements else 'N',
                'Y' if a.yellow_handrails else 'N',
                'Y' if a.priority_seating else 'N',
                'Y' if a.contrasting_seat_covers else 'N',
                'Y' if a.induction_loop else 'N',
            ])
    autowidth(ws5)

    # Sheet 6 — Declarations
    ws6 = wb.create_sheet('Declarations')
    ws6.append(['Operator Name', 'Data Year', 'Signed By', 'Printed Name', 'Position', 'Date'])
    style_header(ws6)
    for r in returns:
        if hasattr(r, 'declaration') and r.declaration:
            d = r.declaration
            ws6.append([
                r.operator_name, r.data_year,
                d.signed_by, d.printed_name, d.position_in_company,
                d.declaration_date.strftime('%d/%m/%Y') if d.declaration_date else '',
            ])
    autowidth(ws6)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="BORS_All_Operators.xlsx"'
    wb.save(response)
    return response


@staff_member_required(login_url='/admin/login/')
def export_all_json(request):
    import json
    from django.core.serializers.json import DjangoJSONEncoder

    returns = (
        OperatorReturn.objects
        .prefetch_related('licences', 'emissions__route_usages', 'accessibility', 'declaration')
        .order_by('operator_name', 'data_year')
    )

    data = []
    for r in returns:
        decl = getattr(r, 'declaration', None)
        data.append({
            'operator_name': r.operator_name,
            'trading_name': r.trading_name,
            'data_year': r.data_year,
            'county': r.county,
            'annual_revenue': str(r.annual_revenue),
            'taxsaver': r.get_taxsaver_display(),
            'taxsaver_other': r.taxsaver_other,
            'status': r.status,
            'submitted_at': r.submitted_at.isoformat() if r.submitted_at else None,
            'licences': [
                {
                    'route_no': lic.route_no,
                    'licence_no': lic.licence_no,
                    'on_free_travel': lic.on_free_travel,
                    'on_yasc': lic.on_yasc,
                    'date_joined': lic.date_joined.isoformat() if lic.date_joined else None,
                    'service_commenced': lic.service_commenced,
                    'service_ceased': lic.service_ceased,
                    'date_ceased': lic.date_ceased.isoformat() if lic.date_ceased else None,
                    'total_passenger_journeys': lic.total_passenger_journeys,
                    'free_travel_percent': str(lic.free_travel_percent),
                    'km_operated': str(lic.km_operated),
                    'daily_pvr': lic.daily_pvr,
                }
                for lic in r.licences.all()
            ],
            'emissions': [
                {
                    'res_id': v.res_id,
                    'vehicle_reg': v.vehicle_reg,
                    'regularly_deployed': v.regularly_deployed,
                    'make': v.make,
                    'model_version': v.model_version,
                    'transmission': v.transmission,
                    'engine_type': v.engine_type,
                    'fuel_type': v.fuel_type,
                    'euro_standard': v.euro_standard,
                    'make_model_correct': v.make_model_correct,
                    'seats_on_record': v.seats_on_record,
                    'correct_seats': v.correct_seats,
                    'avl_gps': v.avl_gps,
                    'route_usage': [
                        {'route_no': u.route_no, 'usage_percent': str(u.usage_percent) if u.usage_percent is not None else None}
                        for u in v.route_usages.all()
                    ],
                }
                for v in r.emissions.all()
            ],
            'accessibility': [
                {
                    'res_id': a.res_id,
                    'vehicle_reg': a.vehicle_reg,
                    'wheelchair_spaces': a.wheelchair_spaces,
                    'wheelchair_access_type': a.wheelchair_access_type,
                    'mechanical_ramp': a.mechanical_ramp,
                    'centre_door_ramps': a.centre_door_ramps,
                    'exterior_displays': a.exterior_displays,
                    'interior_stop_displays': a.interior_stop_displays,
                    'audio_announcements': a.audio_announcements,
                    'yellow_handrails': a.yellow_handrails,
                    'priority_seating': a.priority_seating,
                    'contrasting_seat_covers': a.contrasting_seat_covers,
                    'induction_loop': a.induction_loop,
                }
                for a in r.accessibility.all()
            ],
            'declaration': {
                'signed_by': decl.signed_by,
                'printed_name': decl.printed_name,
                'position_in_company': decl.position_in_company,
                'declaration_date': decl.declaration_date.isoformat() if decl.declaration_date else None,
            } if decl else None,
        })

    response = HttpResponse(
        json.dumps(data, cls=DjangoJSONEncoder, indent=2),
        content_type='application/json',
    )
    response['Content-Disposition'] = 'attachment; filename="BORS_All_Operators.json"'
    return response


# ── Seed data import ──────────────────────────────────────────────────────────

@staff_member_required(login_url='/admin/login/')
def admin_seed_import(request):
    result = None
    if request.method == 'POST' and request.FILES.get('seed_file'):
        f = request.FILES['seed_file']
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            messages.error(request, f'Could not open file: {e}')
            return redirect('admin_seed_import')

        stats = {'operators_added': 0, 'operators_skipped': 0,
                 'licences_added': 0, 'licences_skipped': 0,
                 'vehicles_added': 0, 'vehicles_updated': 0, 'vehicles_skipped': 0,
                 'errors': []}

        if 'Operators' in wb.sheetnames:
            for row in wb['Operators'].iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name  = str(row[0]).strip()
                email = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not name or not email:
                    stats['operators_skipped'] += 1
                    continue
                try:
                    OperatorAccess.objects.get(operator_name__iexact=name)
                    stats['operators_skipped'] += 1
                except OperatorAccess.DoesNotExist:
                    OperatorAccess.objects.create(operator_name=name, operator_email=email)
                    stats['operators_added'] += 1

        if 'Licences' in wb.sheetnames:
            for row in wb['Licences'].iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                op_name  = str(row[0]).strip()
                route_no = str(row[1]).strip() if row[1] else ''
                if not route_no:
                    stats['licences_skipped'] += 1
                    continue
                try:
                    access = OperatorAccess.objects.get(operator_name__iexact=op_name)
                    _, created = SeedLicence.objects.get_or_create(
                        operator_access=access, route_no=route_no
                    )
                    stats['licences_added' if created else 'licences_skipped'] += 1
                except OperatorAccess.DoesNotExist:
                    stats['errors'].append(f'Operator not found: "{op_name}"')

        if 'Vehicles' in wb.sheetnames:
            for row in wb['Vehicles'].iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                op_name     = str(row[0]).strip()
                res_id      = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                vehicle_reg = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                make        = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                model_ver   = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
                trans       = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
                engine      = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ''
                seats       = row[7]              if len(row) > 7 and row[7] is not None else None
                if not vehicle_reg:
                    stats['vehicles_skipped'] += 1
                    continue
                try:
                    access = OperatorAccess.objects.get(operator_name__iexact=op_name)
                    try:
                        seats_int = int(seats) if seats is not None else None
                    except (ValueError, TypeError):
                        seats_int = None
                    _, created = SeedVehicle.objects.update_or_create(
                        operator_access=access, vehicle_reg=vehicle_reg,
                        defaults={'res_id': res_id, 'make': make, 'model_version': model_ver,
                                  'transmission': trans, 'engine_type': engine,
                                  'seats_on_record': seats_int},
                    )
                    stats['vehicles_added' if created else 'vehicles_updated'] += 1
                except OperatorAccess.DoesNotExist:
                    stats['errors'].append(f'Operator not found: "{op_name}"')

        result = stats
        if not stats['errors']:
            messages.success(request, 'Seed file imported successfully.')
        else:
            messages.warning(request, f'{len(stats["errors"])} operator(s) not matched — see details below.')

    accesses = OperatorAccess.objects.prefetch_related(
        'seed_licences', 'seed_vehicles'
    ).order_by('operator_name')

    return render(request, 'returns/admin_seed_import.html', {
        'accesses': accesses,
        'result': result,
    })


@staff_member_required(login_url='/admin/login/')
def download_seed_template(request):
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')

    def style(ws):
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hfill

    ws0 = wb.active
    ws0.title = 'Operators'
    ws0.append(['Operator Name', 'Email'])
    style(ws0)
    ws0.column_dimensions['A'].width = 35
    ws0.column_dimensions['B'].width = 35

    ws1 = wb.create_sheet('Licences')
    ws1.append(['Operator Name', 'Route No'])
    style(ws1)
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet('Vehicles')
    ws2.append(['Operator Name', 'RES ID', 'Vehicle Reg No', 'Make',
                'Model Version', 'Transmission', 'Engine Type', 'Seats on Record'])
    style(ws2)
    for col, w in zip('ABCDEFGH', [35, 10, 18, 18, 25, 18, 18, 16]):
        ws2.column_dimensions[col].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="BORS_Seed_Template.xlsx"'
    wb.save(response)
    return response

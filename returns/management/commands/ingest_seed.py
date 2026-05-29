"""
Management command: python manage.py ingest_seed [--file path/to/file.xlsx]

Reads every .xlsx in seed_data/ (or the file given with --file) and
upserts OperatorAccess, SeedLicence, and SeedVehicle records.

Excel format
------------
Sheet "Operators" : Operator Name | Email
Sheet "Licences"  : Operator Name | Route No
Sheet "Vehicles"  : Operator Name | RES ID | Vehicle Reg No | Make | Model Version | Transmission | Engine Type | Seats on Record
"""

import os
import glob as _glob

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from returns.models import OperatorAccess, SeedLicence, SeedVehicle


class Command(BaseCommand):
    help = 'Ingest NTA seed data (routes + vehicles) from Excel files in seed_data/'

    def add_arguments(self, parser):
        parser.add_argument('--file', dest='file', default=None,
                            help='Path to a specific Excel file to ingest.')

    def handle(self, *args, **options):
        if options['file']:
            files = [options['file']]
        else:
            base = os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', 'seed_data')
            base = os.path.abspath(base)
            files = _glob.glob(os.path.join(base, '*.xlsx'))
            if not files:
                raise CommandError(f'No .xlsx files found in {base}')

        for path in files:
            self.stdout.write(f'Processing {path}...')
            self._process(path)

        self.stdout.write(self.style.SUCCESS('Done.'))

    def _process(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise CommandError(f'Cannot open {path}: {e}')

        if 'Operators' in wb.sheetnames:
            self._ingest_operators(wb['Operators'])
        else:
            self.stdout.write(self.style.WARNING('  No "Operators" sheet found — skipping operators.'))

        if 'Licences' in wb.sheetnames:
            self._ingest_licences(wb['Licences'])
        else:
            self.stdout.write(self.style.WARNING('  No "Licences" sheet found — skipping licences.'))

        if 'Vehicles' in wb.sheetnames:
            self._ingest_vehicles(wb['Vehicles'])
        else:
            self.stdout.write(self.style.WARNING('  No "Vehicles" sheet found — skipping vehicles.'))

    def _ingest_operators(self, ws):
        added = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name  = str(row[0]).strip()
            email = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not name or not email:
                skipped += 1
                continue
            try:
                OperatorAccess.objects.get(operator_name__iexact=name)
                skipped += 1
            except OperatorAccess.DoesNotExist:
                OperatorAccess.objects.create(operator_name=name, operator_email=email)
                added += 1
        self.stdout.write(f'  Operators: {added} created, {skipped} already existed / skipped.')

    def _get_access(self, operator_name):
        try:
            return OperatorAccess.objects.get(operator_name__iexact=operator_name.strip())
        except OperatorAccess.DoesNotExist:
            self.stdout.write(self.style.WARNING(
                f'    No OperatorAccess found for "{operator_name}" — row skipped.'
            ))
            return None
        except OperatorAccess.MultipleObjectsReturned:
            self.stdout.write(self.style.WARNING(
                f'    Multiple matches for "{operator_name}" — row skipped.'
            ))
            return None

    def _ingest_licences(self, ws):
        added = updated = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            operator_name = str(row[0]).strip()
            route_no = str(row[1]).strip() if row[1] else ''
            if not route_no:
                skipped += 1
                continue
            access = self._get_access(operator_name)
            if not access:
                skipped += 1
                continue
            _, created = SeedLicence.objects.get_or_create(
                operator_access=access, route_no=route_no
            )
            if created:
                added += 1
            else:
                updated += 1
        self.stdout.write(f'  Licences: {added} added, {updated} already existed, {skipped} skipped.')

    def _ingest_vehicles(self, ws):
        added = updated = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            operator_name = str(row[0]).strip()
            res_id       = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            vehicle_reg  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            make         = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            model_ver    = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            transmission = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            engine_type  = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ''
            seats        = row[7] if len(row) > 7 and row[7] is not None else None

            if not vehicle_reg:
                skipped += 1
                continue
            access = self._get_access(operator_name)
            if not access:
                skipped += 1
                continue

            try:
                seats_int = int(seats) if seats is not None else None
            except (ValueError, TypeError):
                seats_int = None

            obj, created = SeedVehicle.objects.update_or_create(
                operator_access=access, vehicle_reg=vehicle_reg,
                defaults={
                    'res_id': res_id,
                    'make': make,
                    'model_version': model_ver,
                    'transmission': transmission,
                    'engine_type': engine_type,
                    'seats_on_record': seats_int,
                },
            )
            if created:
                added += 1
            else:
                updated += 1
        self.stdout.write(f'  Vehicles: {added} added, {updated} updated, {skipped} skipped.')
